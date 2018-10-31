import openpyxl
from .constants import *
import mimetypes
import operator


class TrackingSheet:

    def __init__(self, file_name, lst, school_name):
        self.filename = file_name
        self.lst = lst
        self.school_name = school_name
        mimetypes.add_type('image/wmf', '.wmf')

    def export(self, learners):
        # sort alphabetically
        learners.sort(key=lambda x: (x.grade, x.clss, x.surname))

        # open workbook
        srcfile = openpyxl.load_workbook(self.filename, read_only=False, keep_vba=True)
        sheet = srcfile[SHEET_TRACKING_INFO]

        # write teacher and school name
        sheet[REF_LST_NAME] = "LST: " + self.lst
        sheet[REF_SCHOOL_NAME] = self.school_name

        # write learners
        i = 0
        ROW_END = ROW_START + len(learners)
        for row in range(ROW_START, ROW_END):
            l_data = (learners[i]).get_dict()

            # set nr
            sheet.cell(row=row, column=1).value = i + 1

            for col in REF_COLS:
                sheet.cell(row=row, column=col).value = l_data[col]

            i += 1

        out = self.filename.split('/')[-1]
        srcfile.save('../out/' + out)